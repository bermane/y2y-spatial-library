"""Management of queue/processing/pending.xlsx — the Phase-1 review sheet.

The sheet is persistent across scans (DESIGN.md §8). On each scan the
pipeline appends rows for newly-staged datasets; on approve it removes
rows that have been promoted. Rows are identified by ``dataset_id``.

Column layout: ``ready`` first (the steward's control), then every
inventory column in canonical order (see ``inventory_manager``), then
``_validation_error`` last as a pipeline-written error sink.

Column roles (colour-coded in the header):

    control      ``ready``                  — steward flips to TRUE to promote
    locked       auto-filled, never edit    — intrinsic snapshot + id + paths
    overridable  auto-filled, can edit      — best-guess fields (title, category…)
    required     empty, must be filled      — extrinsic Dublin Core+ metadata
    optional     empty, may be filled       — AGOL item id, notes
    error        pipeline-written           — ``_validation_error`` column

See DESIGN.md §7 for the full inventory schema and §8 for the two-phase
ingestion workflow.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .inventory_manager import assert_not_locked

PENDING_FILENAME = "pending.xlsx"
PENDING_SHEET_NAME = "pending"

READY_COLUMN = "ready"
ERROR_COLUMN = "_validation_error"

# Pipeline-only column that lives in pending.xlsx but never makes it
# into inventory.xlsx. ``target_filename`` is the transformation input
# the steward confirms (or overrides) before approval.
TARGET_FILENAME_COLUMN = "target_filename"


# Pending-sheet column order is *not* the inventory column order — it's
# tuned for the steward's review workflow:
#
#   1. control + error sink so the row's promote-state is visible at a glance
#   2. source identity (what arrived) so the steward identifies the row
#   3. transform inputs (what the pipeline will do with it)
#   4. categorisation (auto-prefilled where possible; steward confirms/overrides)
#   5. required extrinsic metadata (where the editing actually happens)
#   6. optional / freeform
#   7. derived / locked / will-be-populated-at-approve (tail; ignore during review)
#
# inventory.xlsx keeps the canonical-grouping order from inventory_manager;
# this sheet's order is steward-facing only.
COLUMNS: list[tuple[str, str]] = [
    # 1. control + error
    (READY_COLUMN, "control"),
    (ERROR_COLUMN, "error"),
    # 2. identity & source provenance
    ("dataset_id", "locked"),
    ("source_format", "locked"),
    ("source_filename", "locked"),
    ("source_crs", "locked"),
    ("source_layer", "locked"),
    # 3. transform inputs
    (TARGET_FILENAME_COLUMN, "transform"),
    ("classification", "overridable"),
    # 4. categorisation (auto-prefilled where possible)
    ("category", "overridable"),
    ("subcategory", "overridable"),
    ("status", "overridable"),
    # 5. required extrinsic metadata
    ("title", "required"),
    ("summary", "required"),
    ("description", "required"),
    ("tags", "required"),
    ("terms_of_use", "required"),
    ("acknowledgements", "required"),
    # AGOL publish target sits with the steward-authored metadata
    # block (immediately after acknowledgements, per steward
    # preference 2026-05-27). Auto-prefilled from format:
    # ``feature-layer`` for vectors, ``imagery-layer`` for rasters.
    # Steward overrides to ``vector-tile-layer`` for large vectors
    # that should be delivered as cached tiles. See DESIGN.md §15.
    ("agol_format", "overridable"),
    ("data_steward", "required"),
    # 6. optional
    ("agol_item_id", "optional"),
    ("notes", "optional"),
    # 7. derived / locked / will be populated at approve
    ("format", "locked"),
    ("file_path", "locked"),
    ("crs", "locked"),
    ("checksum_sha256", "locked"),
    ("size_bytes", "locked"),
    ("mtime", "locked"),
    ("geographic_extent_bbox", "locked"),
    ("date_added", "locked"),
    ("date_modified", "locked"),
]
COLUMN_NAMES: list[str] = [name for name, _ in COLUMNS]

_FILLS: dict[str, PatternFill] = {
    "control": PatternFill("solid", start_color="FFE066"),      # yellow
    "locked": PatternFill("solid", start_color="D3D3D3"),       # grey
    "overridable": PatternFill("solid", start_color="CFE2F3"),  # light blue
    "transform": PatternFill("solid", start_color="FCE5CD"),    # light orange
    "required": PatternFill("solid", start_color="F4CCCC"),     # pink
    "optional": PatternFill("solid", start_color="D9EAD3"),     # light green
    "error": PatternFill("solid", start_color="CC0000"),        # dark red
}

_HEADER_FONT_DARK = Font(bold=True, color="000000")
_HEADER_FONT_LIGHT = Font(bold=True, color="FFFFFF")


def load_pending(path: Path) -> list[dict[str, Any]]:
    """Return rows as list-of-dicts. Empty list if the file doesn't exist."""
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb[PENDING_SHEET_NAME]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def save_pending(path: Path, rows: list[dict[str, Any]]) -> None:
    """Write ``rows`` to ``path``, replacing any existing content.

    Refuses to write if Excel has the file open (race-prevention).
    """
    assert_not_locked(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = PENDING_SHEET_NAME

    for col_idx, (name, role) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT_LIGHT if role == "error" else _HEADER_FONT_DARK
        cell.fill = _FILLS[role]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(COLUMN_NAMES, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))

    for col_idx, name in enumerate(COLUMN_NAMES, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(12, min(40, len(name) + 4))

    ws.freeze_panes = "A2"
    wb.save(path)


def append_pending(path: Path, new_rows: list[dict[str, Any]]) -> None:
    """Load the existing sheet (if any), append new rows, save.

    Rows whose ``dataset_id`` already appears in the sheet are skipped,
    so re-running scan on the same file is idempotent.
    """
    existing = load_pending(path)
    seen = {r.get("dataset_id") for r in existing if r.get("dataset_id")}
    merged = list(existing)
    for row in new_rows:
        if row.get("dataset_id") in seen:
            continue
        merged.append(row)
    save_pending(path, merged)


def delete_if_empty(path: Path) -> bool:
    """Delete pending.xlsx if it has no data rows. Returns True if deleted."""
    if not path.exists():
        return False
    rows = load_pending(path)
    if not rows:
        path.unlink()
        return True
    return False
