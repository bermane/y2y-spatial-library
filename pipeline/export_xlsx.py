"""Generate ``inventory/inventory.xlsx`` from the SQLite catalogue.

The xlsx is a **regenerated, read-only artifact** — a steward-friendly
view of ``inventory.db``. Editing the xlsx changes nothing in the
catalogue. Re-export overwrites it.

Layout choices:

* Column order roughly follows the legacy xlsx grouping (identity →
  extrinsic → lifecycle → location → intrinsic → spatial → AGOL →
  source) so the steward's muscle memory carries over, but every
  schema column is included so the export is a faithful render.
* Header colour-coding still distinguishes "auto-filled snapshot",
  "categorical", "freeform", etc. — purely informational since the
  whole sheet is read-only now.
* A second sheet ``changelog`` exposes the audit log so the whole
  catalogue is browsable in one workbook.
* ``inventory.xlsx`` is rejected if Excel has it open (lock file
  present); the steward must close it before re-exporting.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import inventory_manager

INVENTORY_SHEET_NAME = "inventory"
CHANGELOG_SHEET_NAME = "changelog"

# (column_name, role) — roles drive header colour. Roles preserved
# from the legacy xlsx where applicable; new schema columns get
# "snapshot" or "agol" or "computed" as appropriate.
_INVENTORY_LAYOUT: list[tuple[str, str]] = [
    # identity & taxonomy
    ("dataset_id", "locked"),
    ("dataset_type", "locked"),
    ("category", "overridable"),
    ("subcategory", "overridable"),
    # extrinsic metadata (steward-authored)
    ("title", "required"),
    ("summary", "required"),
    ("description", "required"),
    ("tags", "required"),
    ("terms_of_use", "required"),
    ("acknowledgements", "required"),
    ("data_steward", "required"),
    ("internal_notes", "optional"),
    # lifecycle
    ("classification", "overridable"),
    ("status", "overridable"),
    ("date_added", "locked"),
    ("date_modified", "locked"),
    # canonical location
    ("file_path", "locked"),
    ("format", "locked"),
    # intrinsic snapshot
    ("crs", "locked"),
    ("checksum_sha256", "locked"),
    ("size_bytes", "locked"),
    ("mtime", "locked"),
    ("geographic_extent_bbox", "locked"),
    ("footprint_wkt", "locked"),
    # spatial properties
    ("temporal_start", "optional"),
    ("temporal_end", "optional"),
    ("feature_count", "locked"),
    ("raster_width", "locked"),
    ("raster_height", "locked"),
    ("pixel_size_x", "locked"),
    ("pixel_size_y", "locked"),
    # AGOL linkage (reserved)
    ("agol_item_id", "agol"),
    ("agol_published_at", "agol"),
    ("last_synced_at", "agol"),
    ("sync_status", "agol"),
    # source provenance
    ("source_format", "locked"),
    ("source_filename", "locked"),
    ("source_crs", "locked"),
    ("source_layer", "locked"),
]
_INVENTORY_COLUMN_NAMES: list[str] = [c for c, _ in _INVENTORY_LAYOUT]

_CHANGELOG_LAYOUT: list[tuple[str, str]] = [
    ("id", "locked"),
    ("timestamp", "locked"),
    ("dataset_id", "locked"),
    ("action", "locked"),
    ("field_changed", "locked"),
    ("old_value", "locked"),
    ("new_value", "locked"),
    ("note", "locked"),
    ("actor", "locked"),
]
_CHANGELOG_COLUMN_NAMES: list[str] = [c for c, _ in _CHANGELOG_LAYOUT]

_FILLS: dict[str, PatternFill] = {
    "locked": PatternFill("solid", start_color="D3D3D3"),       # grey
    "overridable": PatternFill("solid", start_color="CFE2F3"),  # light blue
    "required": PatternFill("solid", start_color="F4CCCC"),     # pink
    "optional": PatternFill("solid", start_color="D9EAD3"),     # light green
    "agol": PatternFill("solid", start_color="EAD1DC"),         # mauve (reserved)
}

_HEADER_FONT = Font(bold=True, color="000000")


def _safety_check(out_path: Path) -> None:
    """Refuse to write if Excel has the target xlsx open (lock-file race)."""
    if out_path.exists():
        inventory_manager.assert_not_locked(out_path)


def export(
    db_path: Path,
    out_path: Path,
) -> tuple[int, int]:
    """Write a fresh xlsx to ``out_path`` from the catalogue at ``db_path``.

    Returns ``(dataset_count, changelog_count)``.
    """
    _safety_check(out_path)

    inventory_rows = inventory_manager.load_inventory(db_path)
    changelog_rows = inventory_manager.load_changelog(db_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Default sheet → inventory.
    ws_inv = wb.active
    ws_inv.title = INVENTORY_SHEET_NAME
    _write_sheet(ws_inv, _INVENTORY_LAYOUT, _INVENTORY_COLUMN_NAMES, inventory_rows)

    ws_log = wb.create_sheet(CHANGELOG_SHEET_NAME)
    _write_sheet(ws_log, _CHANGELOG_LAYOUT, _CHANGELOG_COLUMN_NAMES, changelog_rows)

    wb.save(out_path)
    return len(inventory_rows), len(changelog_rows)


def _write_sheet(
    ws: Any,
    layout: list[tuple[str, str]],
    column_names: list[str],
    rows: list[dict[str, Any]],
) -> None:
    """Render header (with role-coloured fill) + body for one sheet."""
    for col_idx, (name, role) in enumerate(layout, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _FILLS.get(role, _FILLS["locked"])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(column_names, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))

    for col_idx, name in enumerate(column_names, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(12, min(40, len(name) + 4))

    ws.freeze_panes = "A2"
