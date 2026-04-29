"""Inventory manager: read/write inventory.xlsx and append to changelog.md.

The inventory is the canonical record of metadata and history for every
dataset in ``library/``. The filesystem is canonical for *existence*
and *location*; the inventory is canonical for everything else
(provenance, license, checksum, version, stewardship, AGOL linkage).

Column order and groupings come from DESIGN.md §7. Changelog format
comes from inventory/changelog.md and is append-only — never edit or
regenerate past entries.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

INVENTORY_FILENAME = "inventory.xlsx"
INVENTORY_SHEET_NAME = "inventory"
CHANGELOG_FILENAME = "changelog.md"


class InventoryLockedError(RuntimeError):
    """Raised when an .xlsx pipeline-write target appears to be open in Excel.

    Excel writes a ``~$<filename>`` lock file in the same directory while
    a workbook is open. If the pipeline writes to that file at the same
    time, Excel's subsequent save will silently overwrite the pipeline's
    changes when the user saves manually. Refusing to write when a lock
    file is present prevents this race.
    """


def assert_not_locked(path: Path) -> None:
    """Raise :class:`InventoryLockedError` if ``path`` appears open in Excel.

    Detection: presence of a sibling ``~$<filename>`` lock file. Excel
    creates these when a workbook opens and removes them when it closes
    (or sometimes leaves them if it crashed). False positives are
    cheap to fix — just delete the stray lock file.
    """
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        raise InventoryLockedError(
            f"{path.name} appears to be open in Excel ({lock.name} present in "
            f"{path.parent}). Close Excel before running this command — "
            f"otherwise your manual save can clobber the pipeline's. "
            f"If no Excel instance is actually open, delete the stray lock: "
            f"`rm '{lock}'`"
        )


# Ordered (name, role) — matches DESIGN.md §7 groupings.
# Roles: "locked" (auto-filled, never edited), "overridable" (auto-filled
# but the steward may correct), "required" (steward must fill), "optional".
# Column order is tuned for steward-UX readability when scanning
# inventory.xlsx, not for canonical-grouping symmetry. The conceptual
# groupings in DESIGN.md §7 are still the source-of-truth for
# *what* each column means; this list controls *how* they're laid
# out in the workbook.
INVENTORY_COLUMNS: list[tuple[str, str]] = [
    # Identity (lead): which dataset is this and where does it sit in the taxonomy?
    ("dataset_id", "locked"),
    ("category", "overridable"),
    ("subcategory", "overridable"),
    # AGOL-facing extrinsic content: what the steward authored, plus
    # AGOL linkage and freeform notes (kept together so the whole AGOL
    # block is contiguous).
    ("title", "required"),
    ("summary", "required"),
    ("description", "required"),
    ("tags", "required"),
    ("terms_of_use", "required"),
    ("acknowledgements", "required"),
    ("data_steward", "required"),
    ("agol_item_id", "optional"),
    ("notes", "optional"),
    # Type and lifecycle state.
    ("classification", "overridable"),
    ("status", "overridable"),
    # Dates and on-disk location of the canonical file.
    ("date_added", "locked"),
    ("date_modified", "locked"),
    ("file_path", "locked"),
    ("format", "locked"),
    # Intrinsic snapshot (drift-detection only).
    ("crs", "locked"),
    ("checksum_sha256", "locked"),
    ("size_bytes", "locked"),
    ("mtime", "locked"),
    ("geographic_extent_bbox", "locked"),
    # Source provenance — pushed to the back: forensic columns the
    # steward reads occasionally, never edits.
    ("source_format", "locked"),
    ("source_filename", "locked"),
    ("source_crs", "locked"),
    ("source_layer", "locked"),
]

INVENTORY_COLUMN_NAMES: list[str] = [name for name, _ in INVENTORY_COLUMNS]

_FILLS: dict[str, PatternFill] = {
    "locked": PatternFill("solid", start_color="D3D3D3"),       # grey
    "overridable": PatternFill("solid", start_color="CFE2F3"),  # light blue
    "required": PatternFill("solid", start_color="F4CCCC"),     # pink
    "optional": PatternFill("solid", start_color="D9EAD3"),     # light green
}

_HEADER_FONT = Font(bold=True, color="000000")


# --- inventory.xlsx I/O --------------------------------------------------

def load_inventory(path: Path) -> list[dict[str, Any]]:
    """Return inventory rows. Empty list if the file does not exist."""
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb[INVENTORY_SHEET_NAME]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def save_inventory(path: Path, rows: list[dict[str, Any]]) -> None:
    """Write rows to ``path``, replacing any existing content.

    Refuses to write if Excel has the file open (race-prevention).
    """
    assert_not_locked(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = INVENTORY_SHEET_NAME

    for col_idx, (name, role) in enumerate(INVENTORY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _FILLS[role]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(INVENTORY_COLUMN_NAMES, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))

    for col_idx, name in enumerate(INVENTORY_COLUMN_NAMES, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(12, min(40, len(name) + 4))

    ws.freeze_panes = "A2"
    wb.save(path)


def append_inventory(path: Path, new_rows: list[dict[str, Any]]) -> None:
    """Append rows to inventory.xlsx (creating it if needed).

    Rows whose ``dataset_id`` already exists in the inventory are
    skipped. Use ``save_inventory`` directly for replacement semantics.
    """
    existing = load_inventory(path)
    seen = {r.get("dataset_id") for r in existing if r.get("dataset_id")}
    merged = list(existing)
    for row in new_rows:
        if row.get("dataset_id") in seen:
            continue
        # Preserve only canonical columns to keep inventory clean
        clean = {k: row.get(k) for k in INVENTORY_COLUMN_NAMES}
        merged.append(clean)
    save_inventory(path, merged)


# --- changelog.md (append-only) -----------------------------------------

def append_changelog(
    changelog_path: Path,
    *,
    timestamp: str,
    action: str,
    dataset_id: str,
    actor: str,
    path: str | None,
    detail: str,
) -> None:
    """Append one entry to the changelog. Never edits prior entries.

    Block format (matches inventory/changelog.md spec):

        ## YYYY-MM-DDTHH:MM:SSZ — <action> — <dataset_id>
        actor:  <data_steward>
        path:   <library-relative path, or "—">
        detail: <one-line summary>
    """
    changelog_path.parent.mkdir(parents=True, exist_ok=True)
    if not changelog_path.exists():
        # Create with header so the file is self-describing if it ever
        # gets written by approve() before the steward seeds it manually.
        changelog_path.write_text(
            "# Y2Y Spatial Library — Changelog\n\n"
            "Append-only audit log. **Never edit past entries. "
            "Never regenerate this file.**\n\n"
        )

    block = (
        f"## {timestamp} — {action} — {dataset_id}\n"
        f"actor:  {actor}\n"
        f"path:   {path or '—'}\n"
        f"detail: {detail}\n\n"
    )
    with changelog_path.open("a", encoding="utf-8") as f:
        f.write(block)
